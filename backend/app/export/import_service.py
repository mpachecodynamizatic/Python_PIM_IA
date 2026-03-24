"""
Excel import service: parse, validate and apply xlsx files.
"""
from __future__ import annotations

import io
import uuid
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.export.configs import ExportConfig, ExportField, get_config, get_fk_model_map
from app.export.validators import coerce_value


# ── Result types ──────────────────────────────────────────────────────────────

@dataclass
class ImportIssue:
    row: int
    field_key: str
    code: str        # "error_*" = blocking  |  "warn_*" = non-blocking
    message: str


@dataclass
class ImportPreviewRow:
    row: int
    mode: str        # "create" | "update"
    data: dict       # upsert_key values for quick identification


@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0


@dataclass
class ImportValidationResult:
    total: int
    valid: int
    errors: list[ImportIssue]
    warnings: list[ImportIssue]
    preview: list[ImportPreviewRow]
    # Kept for direct apply (not serialised)
    _valid_rows: list[dict] = field(default_factory=list, repr=False)

    @property
    def has_blocking_errors(self) -> bool:
        return any(i.code.startswith("error_") for i in self.errors)


# ── Parsing ───────────────────────────────────────────────────────────────────

def _parse_excel_bytes(file_bytes: bytes, config: ExportConfig) -> list[dict]:
    """Parse xlsx bytes into a list of raw row dicts keyed by column *label*."""
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []

    # Build label → key mapping from config
    label_to_key: dict[str, str] = {ef.label: ef.key for ef in config.fields}

    # Detect columns present in the file
    col_keys: list[str | None] = []
    for cell_val in header_row:
        raw = str(cell_val).strip() if cell_val is not None else ""
        col_keys.append(label_to_key.get(raw))  # None if unrecognised

    result = []
    for row in rows_iter:
        row_dict: dict[str, Any] = {}
        for col_idx, key in enumerate(col_keys):
            if key is not None and col_idx < len(row):
                row_dict[key] = row[col_idx]
        if any(v is not None and v != "" for v in row_dict.values()):
            result.append(row_dict)

    wb.close()
    return result


# ── FK batch check ─────────────────────────────────────────────────────────────

async def _batch_fk_check(
    config: ExportConfig,
    rows: list[dict],
    db: AsyncSession,
) -> dict[tuple[str, str], set]:
    """For every FK constraint in the config, query which values exist in the DB.
    Returns {(table, column): {existing_values}}
    """
    fk_model_map = get_fk_model_map()
    # Collect needed values
    needed: dict[tuple[str, str], set] = {}
    for ef in config.fields:
        if ef.fk is None:
            continue
        key = (ef.fk.table, ef.fk.column)
        for row in rows:
            val = row.get(ef.key)
            if val is not None and val != "":
                needed.setdefault(key, set()).add(val)

    fk_exists: dict[tuple[str, str], set] = {}
    for (table, col), values in needed.items():
        model_cls = fk_model_map.get((table, col))
        if model_cls is None:
            continue
        col_attr = getattr(model_cls, col)
        result = await db.execute(select(col_attr).where(col_attr.in_(values)))
        fk_exists[(table, col)] = set(result.scalars().all())

    return fk_exists


# ── Upsert mode detection ─────────────────────────────────────────────────────

async def _get_existing(
    config: ExportConfig,
    row: dict,
    db: AsyncSession,
) -> Any | None:
    """Look up existing row by upsert_key. Returns the ORM object or None."""
    query = select(config.model)
    for uk in config.upsert_key:
        val = row.get(uk)
        if val is None or val == "":
            return None
        col_attr = getattr(config.model, uk, None)
        if col_attr is None:
            return None
        query = query.where(col_attr == val)
    result = await db.execute(query)
    return result.scalar_one_or_none()


# ── Validate ──────────────────────────────────────────────────────────────────

async def validate_import(
    resource: str,
    file_bytes: bytes,
    db: AsyncSession,
) -> ImportValidationResult:
    config = get_config(resource)
    raw_rows = _parse_excel_bytes(file_bytes, config)

    if not raw_rows:
        return ImportValidationResult(
            total=0, valid=0, errors=[], warnings=[], preview=[], _valid_rows=[]
        )

    # Batch FK check
    fk_exists = await _batch_fk_check(config, raw_rows, db)

    errors: list[ImportIssue] = []
    warnings: list[ImportIssue] = []
    preview: list[ImportPreviewRow] = []
    valid_rows: list[dict] = []

    for excel_row_idx, raw in enumerate(raw_rows, start=2):
        row_errors: list[ImportIssue] = []
        row_warnings: list[ImportIssue] = []
        coerced: dict[str, Any] = {}

        for ef in config.fields:
            is_absent = ef.key not in raw  # column not present in this Excel at all
            raw_val = raw.get(ef.key)

            # Readonly fields: accept only for upsert key lookup (skip coercion)
            if ef.readonly:
                if ef.key in config.upsert_key and not is_absent and raw_val not in (None, ""):
                    coerced[ef.key] = str(raw_val).strip()
                continue

            # If the column was not in the Excel at all, only error if required
            if is_absent:
                if ef.required:
                    row_errors.append(ImportIssue(
                        row=excel_row_idx,
                        field_key=ef.key,
                        code="error_missing_column",
                        message=f"[{ef.label}] Columna obligatoria no encontrada en el Excel. Asegurese de que la cabecera contenga '{ef.label}'",
                    ))
                continue  # Skip: model/DB default will be used

            value, err_msg = coerce_value(raw_val, ef)

            if err_msg:
                row_errors.append(ImportIssue(
                    row=excel_row_idx,
                    field_key=ef.key,
                    code="error_invalid_value",
                    message=f"[{ef.label}] {err_msg}",
                ))
                continue

            coerced[ef.key] = value

            # FK check
            if ef.fk and value is not None:
                fk_key = (ef.fk.table, ef.fk.column)
                if fk_key in fk_exists and value not in fk_exists[fk_key]:
                    # Get readable table name
                    table_labels = {
                        "products": "Productos",
                        "categories": "Categorías",
                        "brands": "Marcas",
                        "channels": "Canales",
                        "suppliers": "Proveedores",
                        "attribute_families": "Familias de Atributos",
                        "quality_rule_sets": "Conjuntos de Reglas de Calidad",
                    }
                    table_label = table_labels.get(ef.fk.table, ef.fk.table)
                    row_errors.append(ImportIssue(
                        row=excel_row_idx,
                        field_key=ef.key,
                        code="error_fk_not_found",
                        message=f"[{ef.label}] El valor '{value}' no existe en {table_label}. Verifique que el registro exista antes de importar",
                    ))

        # Upsert mode
        existing = await _get_existing(config, coerced, db)
        mode = "update" if existing is not None else "create"

        if mode == "update":
            # Build identifier for the warning message
            upsert_values = {uk: coerced.get(uk, "?") for uk in config.upsert_key}
            identifier = ", ".join(f"{k}={v}" for k, v in upsert_values.items())
            row_warnings.append(ImportIssue(
                row=excel_row_idx,
                field_key=config.upsert_key[0] if config.upsert_key else "id",
                code="warn_will_overwrite",
                message=f"Se actualizará el registro existente con {identifier}",
            ))

        # Business-rule check: status transitions for products
        if resource == "products" and mode == "update" and existing is not None:
            new_status = coerced.get("status")
            if new_status and new_status != existing.status:
                valid_transitions = {
                    "draft": ["in_review", "ready"],
                    "in_review": ["approved", "draft"],
                    "approved": ["ready", "in_review", "draft"],
                    "ready": ["draft", "retired"],
                    "retired": ["draft"],
                }
                allowed = valid_transitions.get(existing.status, [])
                if new_status not in allowed:
                    row_errors.append(ImportIssue(
                        row=excel_row_idx,
                        field_key="status",
                        code="error_invalid_transition",
                        message=(
                            f"[Estado] Transición inválida de '{existing.status}' a '{new_status}'. "
                            f"Transiciones permitidas desde '{existing.status}': {', '.join(allowed) if allowed else 'ninguna'}"
                        ),
                    ))

        errors.extend(row_errors)
        warnings.extend(row_warnings)

        if not row_errors:
            coerced["_mode"] = mode
            if existing is not None:
                coerced["_existing_pk"] = _get_pk_value(config, existing)
            valid_rows.append(coerced)

        # Preview (first 10 rows regardless of errors)
        preview_data = {uk: coerced.get(uk) for uk in config.upsert_key}
        if len(preview) < 10:
            preview.append(ImportPreviewRow(
                row=excel_row_idx,
                mode=mode,
                data=preview_data,
            ))

    return ImportValidationResult(
        total=len(raw_rows),
        valid=len(valid_rows),
        errors=errors,
        warnings=warnings,
        preview=preview,
        _valid_rows=valid_rows,
    )


def _get_pk_value(config: ExportConfig, obj: Any) -> Any:
    """Return the primary-key value of an ORM object."""
    if not config.auto_pk and config.upsert_key:
        return getattr(obj, config.upsert_key[0], None)
    return getattr(obj, "id", None)


# ── Apply ─────────────────────────────────────────────────────────────────────

async def apply_import(
    resource: str,
    validation: ImportValidationResult,
    actor: str,
    db: AsyncSession,
) -> ImportResult:
    """Apply validated rows to the DB.  Must only be called when
    validation.has_blocking_errors is False.
    """
    from app.models.audit import AuditLog

    config = get_config(resource)
    result = ImportResult()

    for row in validation._valid_rows:
        mode = row.get("_mode", "create")
        existing_pk = row.get("_existing_pk")

        # Strip internal markers
        data = {k: v for k, v in row.items() if not k.startswith("_")}

        if mode == "update" and existing_pk is not None:
            existing = await _fetch_existing_by_pk(config, existing_pk, db)
            if existing is None:
                result.skipped += 1
                continue
            before_state = _obj_to_dict(config, existing)
            _update_obj(config, existing, data)
            db.add(existing)
            await db.flush()
            await db.refresh(existing)
            after_state = _obj_to_dict(config, existing)
            result.updated += 1
        else:
            existing = None
            before_state = None
            new_obj = await _create_obj(config, data, db)
            db.add(new_obj)
            await db.flush()
            await db.refresh(new_obj)
            after_state = _obj_to_dict(config, new_obj)
            result.created += 1

        # Audit entry
        audit = AuditLog(
            resource=resource,
            resource_id=str(after_state.get(config.upsert_key[0], "") if config.upsert_key else ""),
            actor=actor,
            action="import",
            before=before_state,
            after=after_state,
        )
        db.add(audit)

    await db.commit()
    return result


async def _fetch_existing_by_pk(
    config: ExportConfig,
    pk_value: Any,
    db: AsyncSession,
) -> Any | None:
    pk_col = "id" if config.auto_pk else config.upsert_key[0]
    col_attr = getattr(config.model, pk_col, None)
    if col_attr is None:
        return None
    result = await db.execute(select(config.model).where(col_attr == pk_value))
    return result.scalar_one_or_none()


def _update_obj(config: ExportConfig, obj: Any, data: dict) -> None:
    """Update writable fields on an existing ORM object."""
    readonly_keys = {ef.key for ef in config.fields if ef.readonly}
    for key, val in data.items():
        if key in readonly_keys:
            continue
        if hasattr(obj, key):
            setattr(obj, key, val)


async def _create_obj(config: ExportConfig, data: dict, db: AsyncSession) -> Any:
    """Create a new ORM instance."""
    constructor_data: dict[str, Any] = {}
    readonly_keys = {ef.key for ef in config.fields if ef.readonly}

    for ef in config.fields:
        if ef.key in readonly_keys and ef.key not in config.upsert_key:
            continue  # skip auto fields unless they ARE the identity key
        if ef.key in data:
            constructor_data[ef.key] = data[ef.key]

    # Auto-generate primary key where needed
    if config.auto_pk:
        constructor_data["id"] = str(uuid.uuid4())

    # Special handling: users need a hashed_password on create
    if config.resource == "users" and "hashed_password" not in constructor_data:
        from app.core.security import hash_password
        constructor_data["hashed_password"] = hash_password("PIM@ChangeMe!")

    return config.model(**constructor_data)


def _obj_to_dict(config: ExportConfig, obj: Any) -> dict:
    """Serialise an ORM object to a plain dict (for audit logging)."""
    result = {}
    for ef in config.fields:
        val = getattr(obj, ef.key, None)
        if hasattr(val, "isoformat"):
            result[ef.key] = val.isoformat()
        elif isinstance(val, (dict, list)):
            result[ef.key] = val
        else:
            result[ef.key] = val
    return result
