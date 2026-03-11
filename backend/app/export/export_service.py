"""
Excel export service: fetch rows from DB and build an xlsx workbook.
"""
from __future__ import annotations

import io
import json
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.export.configs import ExportConfig, ExportField, get_config
from app.export.validators import serialize_value

# ── Colours ──────────────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_READONLY_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_REQUIRED_FONT = Font(bold=True, size=10)
_EXAMPLE_FONT = Font(italic=True, color="888888", size=9)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _effective_fields(config: ExportConfig, selected_keys: list[str] | None) -> list[ExportField]:
    """Return fields to include, honouring the selected_keys filter."""
    if selected_keys:
        key_set = set(selected_keys)
        return [f for f in config.fields if f.key in key_set]
    return [f for f in config.fields if f.default_include]


def _get_attr(obj: Any, key: str) -> Any:
    """Safely get an attribute; falls back to None."""
    return getattr(obj, key, None)


def _write_header_row(ws: Any, fields: list[ExportField]) -> None:
    for col_idx, ef in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=col_idx, value=ef.label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL if not ef.readonly else _READONLY_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(ef.label) + 4)


# ── Main public API ───────────────────────────────────────────────────────────

async def export_to_excel(
    resource: str,
    selected_keys: list[str] | None,
    filters: dict,
    db: AsyncSession,
) -> io.BytesIO:
    """Fetch rows from DB and return an in-memory xlsx file."""
    config = get_config(resource)
    fields = _effective_fields(config, selected_keys)

    rows = await _fetch_rows(config, filters, db)

    wb = Workbook()
    ws = wb.active
    ws.title = config.label[:31]  # Excel sheet name max 31 chars

    _write_header_row(ws, fields)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}1"

    for row_idx, obj in enumerate(rows, start=2):
        for col_idx, ef in enumerate(fields, start=1):
            raw_value = _get_attr(obj, ef.key)
            ws.cell(row=row_idx, column=col_idx, value=serialize_value(raw_value, ef.type))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


async def generate_template(
    resource: str,
    selected_keys: list[str] | None,
) -> io.BytesIO:
    """Return an xlsx template with headers, one example row and a hidden _meta sheet."""
    config = get_config(resource)
    fields = _effective_fields(config, selected_keys)

    wb = Workbook()
    ws = wb.active
    ws.title = config.label[:31]

    _write_header_row(ws, fields)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}1"

    # Example / placeholder row
    example_row = _build_example_row(fields)
    for col_idx, val in enumerate(example_row, start=1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font = _EXAMPLE_FONT

    # Hidden _meta sheet for the importer
    ms = wb.create_sheet("_meta")
    ms.sheet_state = "hidden"
    ms.append(["resource", "field_key", "label", "type", "required", "choices"])
    for ef in fields:
        ms.append([
            config.resource,
            ef.key,
            ef.label,
            ef.type,
            ef.required,
            json.dumps(ef.choices or []),
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _build_example_row(fields: list[ExportField]) -> list[Any]:
    examples: dict[str, str] = {
        "str": "ejemplo",
        "int": "0",
        "float": "1.0",
        "bool": "true",
        "date": "2026-01-01",
        "datetime": "2026-01-01 00:00",
        "json": "{}",
        "enum": "",
    }
    result = []
    for ef in fields:
        if ef.readonly:
            result.append("(auto)")
        elif ef.type == "enum" and ef.choices:
            result.append(ef.choices[0])
        else:
            result.append(examples.get(ef.type, ""))
    return result


async def _fetch_rows(config: ExportConfig, filters: dict, db: AsyncSession) -> list[Any]:
    """Fetch all matching rows from the DB (no pagination).

    Applies simple equality filters for columns that exist on the model.
    """
    query = select(config.model)

    for key, val in (filters or {}).items():
        if val is None or val == "":
            continue
        col = getattr(config.model, key, None)
        if col is not None:
            query = query.where(col == val)

    result = await db.execute(query)
    return list(result.scalars().all())


def list_fields(resource: str) -> list[dict]:
    """Return field metadata for the UI field selector."""
    config = get_config(resource)
    return [
        {
            "key": ef.key,
            "label": ef.label,
            "type": ef.type,
            "required": ef.required,
            "readonly": ef.readonly,
            "default_include": ef.default_include,
            "choices": ef.choices,
        }
        for ef in config.fields
    ]


def list_resources() -> list[dict]:
    """Return the list of available resources with their label."""
    from app.export.configs import get_configs
    return [
        {"resource": key, "label": cfg.label}
        for key, cfg in get_configs().items()
    ]
