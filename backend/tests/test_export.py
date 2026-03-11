"""
Tests for the Excel export engine.
"""
import io
import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture
async def sample_category(client: AsyncClient, auth_headers):
    r = await client.post(
        "/api/v1/taxonomy/categories",
        json={"name": "Test Cat", "slug": "test-cat-export"},
        headers=auth_headers,
    )
    assert r.status_code == 201
    return r.json()


@pytest.fixture
async def sample_products(client: AsyncClient, auth_headers, sample_category):
    """Create 3 products for export tests."""
    skus = ["EXP-A01", "EXP-A02", "EXP-A03"]
    for sku in skus:
        r = await client.post(
            "/api/v1/products",
            json={"sku": sku, "brand": "ExportBrand", "category_id": sample_category["id"]},
            headers=auth_headers,
        )
        assert r.status_code == 201
    return skus


# ── /export/resources ─────────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_list_resources(client: AsyncClient, auth_headers):
    r = await client.get("/api/v1/export/resources", headers=auth_headers)
    assert r.status_code == 200
    resources = {item["resource"] for item in r.json()}
    assert "products" in resources
    assert "categories" in resources
    assert "users" in resources


# ── /export/{resource}/fields ─────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_get_fields_products(client: AsyncClient, auth_headers):
    r = await client.get("/api/v1/export/products/fields", headers=auth_headers)
    assert r.status_code == 200
    fields = r.json()
    keys = {f["key"] for f in fields}
    assert "sku" in keys
    assert "brand" in keys
    assert "status" in keys
    assert "category_id" in keys


@pytest.mark.asyncio
async def test_get_fields_unknown_resource(client: AsyncClient, auth_headers):
    r = await client.get("/api/v1/export/unknown_resource/fields", headers=auth_headers)
    assert r.status_code == 404


# ── POST /export/{resource} ───────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_export_products_all_fields(
    client: AsyncClient, auth_headers, sample_products
):
    r = await client.post(
        "/api/v1/export/products",
        json={"fields": None, "filters": {}},
        headers=auth_headers,
    )
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]

    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    # Header row exists
    headers = [cell.value for cell in ws[1] if cell.value]
    assert "SKU" in headers
    assert "Marca" in headers

    # Data rows for our 3 products (may have more from other tests)
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    skus_in_excel = {row[headers.index("SKU")] for row in data_rows if row[0]}
    for sku in sample_products:
        assert sku in skus_in_excel


@pytest.mark.asyncio
async def test_export_products_selected_fields(
    client: AsyncClient, auth_headers, sample_products
):
    r = await client.post(
        "/api/v1/export/products",
        json={"fields": ["sku", "brand"], "filters": {}},
        headers=auth_headers,
    )
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert "SKU" in headers
    assert "Marca" in headers
    # Only 2 columns
    assert len([h for h in headers if h]) == 2


@pytest.mark.asyncio
async def test_export_products_with_filter(
    client: AsyncClient, auth_headers, sample_products, sample_category
):
    """Export only products in draft state."""
    r = await client.post(
        "/api/v1/export/products",
        json={"fields": ["sku", "status"], "filters": {"status": "draft"}},
        headers=auth_headers,
    )
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    status_col = headers.index("Estado")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # non-empty row
            assert row[status_col] == "draft"


# ── GET /export/{resource}/template ──────────────────────────────────────────

@pytest.mark.asyncio
async def test_download_template(client: AsyncClient, auth_headers):
    r = await client.get("/api/v1/export/products/template", headers=auth_headers)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]

    wb = load_workbook(io.BytesIO(r.content))
    # Main sheet has headers
    ws = wb.active
    headers = [cell.value for cell in ws[1] if cell.value]
    assert "SKU" in headers

    # _meta hidden sheet exists
    assert "_meta" in wb.sheetnames
    meta_ws = wb["_meta"]
    meta_rows = list(meta_ws.iter_rows(min_row=2, values_only=True))
    assert any(row[0] == "products" for row in meta_rows)


@pytest.mark.asyncio
async def test_download_template_selected_fields(client: AsyncClient, auth_headers):
    r = await client.get(
        "/api/v1/export/products/template?fields=sku&fields=brand",
        headers=auth_headers,
    )
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    headers = [cell.value for cell in ws[1] if cell.value]
    assert len(headers) == 2
    assert "SKU" in headers
    assert "Marca" in headers


# ── Auth guard ────────────────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_export_requires_auth(client: AsyncClient):
    r = await client.post("/api/v1/export/products", json={})
    assert r.status_code in (401, 403)
